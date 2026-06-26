#!/usr/bin/env python3
"""
Update data CSVs from 'CS World Cup Bingo 2026.xlsx'.

Outputs:
  data/completed_matches.csv  — group-stage (and future knockout) results
  data/boards_in_play.csv     — which boards are paid/in play
  data/fair_play.csv          — per-team fair play ratings from the FairPlay sheet
"""
import csv
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

XLSX_PATH = Path(__file__).parent / "CS World Cup Bingo 2026.xlsx"
DATA_DIR = Path(__file__).parent

GROUPS = set("ABCDEFGHIJKL")

STAGE_LABELS = {
    "Round of 32": "R32",
    "Round of 16": "R16",
    "Quarter final": "QF",
    "Semi-Final": "SF",
    "Third place": "Third",
    "Final": "Final",
}


# ---------------------------------------------------------------------------
# boards_in_play.csv
# ---------------------------------------------------------------------------

def update_boards_in_play(wb):
    ws = wb["Boards"]
    rows = [["Board", "In Play"]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        board_id = row[0]
        paid = row[3]
        assigned = row[1] != None
        if not (board_id and isinstance(board_id, str) and "-" in board_id):
            continue
        in_play = "TRUE" if assigned is True else "FALSE"
        rows.append([board_id, in_play])

    out = DATA_DIR / "boards_in_play.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    print(f"  wrote {len(rows)-1} boards → {out.name}")


# ---------------------------------------------------------------------------
# fair_play.csv
# ---------------------------------------------------------------------------

def update_fair_play(wb):
    ws = wb["FairPlay"]
    team_ratings = {}

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] in GROUPS:
            continue
        # Data rows have (None, team, rating, None, team, rating, None, team, rating)
        for name_col, rating_col in [(1, 2), (4, 5), (7, 8)]:
            name = row[name_col] if len(row) > name_col else None
            rating = row[rating_col] if len(row) > rating_col else None
            if name and isinstance(name, str) and rating is not None and isinstance(rating, (int, float)):
                team_ratings[name] = int(rating)

    rows = [["team", "fair_play_rating"]] + sorted(team_ratings.items())
    out = DATA_DIR / "fair_play.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    print(f"  wrote {len(rows)-1} teams → {out.name}")


# ---------------------------------------------------------------------------
# completed_matches.csv helpers
# ---------------------------------------------------------------------------

def _parse_h2h_scores(wb):
    """
    Parse the HeadToHeadComp sheet.

    Returns dict mapping frozenset({team1, team2}) -> (goals_for_team1, goals_for_team2)
    where 'team1' is whichever name comes first alphabetically (canonical order not
    important — we re-order to match DailySchedule later).
    """
    ws = wb["HeadToHeadComp"]
    scores = {}
    seen = set()

    group_teams = []
    in_group = False

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        # Group header: col[1] in A-L and col[4] == 'Pts.'
        if row[1] in GROUPS and len(row) > 4 and row[4] == "Pts.":
            group_teams = []
            in_group = True
            continue

        if not in_group:
            continue

        # Team data row: col[3] is team name, col[4] is numeric points
        team = row[3] if len(row) > 3 else None
        pts = row[4] if len(row) > 4 else None
        if not (team and isinstance(team, str) and team.strip() and isinstance(pts, (int, float))):
            if group_teams:  # blank row after group → end of group
                in_group = False
            continue

        team_idx = len(group_teams)
        group_teams.append(team)

        # Score columns start at index 9; col[9+j] is score vs group_teams[j]
        for j in range(4):
            col_idx = 9 + j
            if col_idx >= len(row):
                continue
            score_str = row[col_idx]
            if score_str is None or not isinstance(score_str, str) or " - " not in score_str:
                continue
            if j >= len(group_teams):
                # Opponent not yet recorded in this pass; will be covered from their row
                continue
            opp = group_teams[j]
            pair = frozenset({team, opp})
            if pair in seen:
                continue
            try:
                g_team, g_opp = map(int, score_str.strip().split(" - "))
            except ValueError:
                continue
            scores[pair] = (team, opp, g_team, g_opp)
            seen.add(pair)

    return scores


def _parse_schedule(wb):
    """
    Parse DailySchedule sheet.

    Returns list of dicts with keys: date, stage, group, team1, team2, match_no
    Only includes rows with resolved team names (skips future knockout matches
    where teams are unknown).
    """
    ws = wb["DailySchedule"]
    schedule = []
    current_stage = "Group Stage"

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 6:
            continue

        date_val = row[1]
        team1 = row[3]
        team2 = row[4]
        match_no = row[5]
        code1 = row[6] if len(row) > 6 else None

        # Section header (e.g. 'Round of 32')
        if isinstance(date_val, str) or (date_val is None and isinstance(team1, str) and team2 is None):
            label = team1 if isinstance(team1, str) else None
            if label and label in STAGE_LABELS:
                current_stage = STAGE_LABELS[label]
            continue

        if not isinstance(date_val, datetime):
            continue
        if not (team1 and team2 and isinstance(team1, str) and isinstance(team2, str)):
            continue

        date_str = date_val.strftime("%Y-%m-%d")

        # Derive group from team code (e.g. 'A1' → 'A')
        group = ""
        if isinstance(code1, str) and len(code1) >= 2 and code1[0] in GROUPS:
            group = code1[0]

        schedule.append({
            "date": date_str,
            "stage": current_stage,
            "group": group,
            "team1": team1,
            "team2": team2,
            "match_no": int(match_no) if isinstance(match_no, float) else match_no,
        })

    return schedule


# ---------------------------------------------------------------------------
# completed_matches.csv
# ---------------------------------------------------------------------------

def update_completed_matches(wb):
    h2h = _parse_h2h_scores(wb)
    schedule = _parse_schedule(wb)

    # Build lookup: frozenset({t1, t2}) -> schedule entry
    sched_lookup = {frozenset({m["team1"], m["team2"]}): m for m in schedule}

    header = ["date", "stage", "group", "team1", "team2",
              "team1_goals", "team2_goals", "extra_time", "penalties", "winner"]
    rows = [header]

    # Sort by match_no to preserve chronological order
    completed_pairs = sorted(
        h2h.items(),
        key=lambda kv: sched_lookup.get(kv[0], {}).get("match_no", 9999)
    )

    for pair, (score_t1, score_t2, g1, g2) in completed_pairs:
        meta = sched_lookup.get(pair)
        if meta is None:
            continue  # no schedule entry (shouldn't happen)

        # Orient goals to match the DailySchedule team order
        t1, t2 = meta["team1"], meta["team2"]
        if score_t1 == t1:
            goals1, goals2 = g1, g2
        else:
            goals1, goals2 = g2, g1

        if goals1 > goals2:
            winner = t1
        elif goals2 > goals1:
            winner = t2
        else:
            winner = ""

        rows.append([
            meta["date"],
            meta["stage"],
            meta["group"],
            t1, t2,
            goals1, goals2,
            0, 0,  # extra_time, penalties (group stage: always 0)
            winner,
        ])

    out = DATA_DIR / "completed_matches.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    print(f"  wrote {len(rows)-1} matches → {out.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not XLSX_PATH.exists():
        sys.exit(f"Not found: {XLSX_PATH}")

    print(f"Loading {XLSX_PATH.name} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    update_boards_in_play(wb)
    update_fair_play(wb)
    update_completed_matches(wb)
    print("Done.")


if __name__ == "__main__":
    main()
